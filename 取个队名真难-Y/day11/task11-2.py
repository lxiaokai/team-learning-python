# -*- coding: utf-8 -*-
import xlrd
import xlwt
import openpyxl
from openpyxl import Workbook
if __name__ =="__main__":
    # 用xlrd 和xlwt进行读写
    # 安装第三方模块xlwt和xlrd
    book = xlrd.open_workbook('test1.xlsx')
    sheet1 = book.sheets()[0]
    nrows = sheet1.nrows
    print('表格总行数',nrows)
    ncols = sheet1.ncols
    print('表格总列数',ncols)
    row3_values = sheet1.row_values(2)
    print('第3行值',row3_values)
    col3_values = sheet1.col_values(2)
    print('第3列值',col3_values)
    cell_3_3 = sheet1.cell(2,2).value
    print('第3行第3列的值',cell_3_3)
    # xlwt 写excel
    workbook = xlwt.Workbook()
    worksheet = workbook.add_sheet('one')
    worksheet.write(0,0,'A1data')
    workbook.save('test2.xls') # 存在xlsx打不开

    #####
    # 用openpyxl读取excel
    workbook1 = openpyxl.load_workbook('test1.xlsx')
    worksheet1 = workbook1.get_sheet_by_name('工作表1')
    row3 = [item.value for item in list(worksheet1.rows)[2]]
    print('第3行值',row3)
    col3 = [item.value for item in list(worksheet1.columns)[2]]
    print('第3列值',col3)
    cell_2_3 = worksheet1.cell(row=2,column=3).value
    print('第2行第3列值',cell_2_3)
    max_row = worksheet1.max_row
    print('最大行',max_row)

    # 用openpyxl 写 excel
    workbook1 = Workbook() # 创建一个新的工作表
    print(workbook1)
    sheet1 = workbook1.active
    sheet1.title = 'first sheet'
    # 增加以及删除表
    workbook1.create_sheet(title='second sheet')
    # 向单元格中写入信息
    sheet1['A1'] = 'HELLO WORLD'
    sheet1.cell(row = 2,column = 2).value = '张三'
    workbook1.save('new.xlsx')
    # 为单元格设置样式
    # https://www.cnblogs.com/gy-ph/p/7827151.html