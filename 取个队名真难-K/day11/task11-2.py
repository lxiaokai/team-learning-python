#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# @Time    : 2018/10/25 4:59 PM
# @Author  : liangk
# @Site    : Excel表格读写
# @File    : task11-2.py
# @Software: PyCharm

from openpyxl import load_workbook, Workbook

# 程序主入口
if __name__ == '__main__':

    # 新建表格,用到Workbook
    # 新建了一个新的工作表（只是还没被保存）。
    wb = Workbook()
    # 默认名字Sheet
    print(wb.sheetnames)

    # 活动的sheet页。空的excel表默认的sheet页就叫Sheet，
    # 如果想改名字，可以直接给title属性赋值
    sheet = wb.active
    sheet.title = 'fat_boy'
    sheet['C3'] = 'fat_boy'
    sheet['C1'] = 'fat_boy'
    for i in range(10):
        sheet["A%d" % (i + 1)].value = i + 1

    # 新建一个工作表,可以指定索引,下标从零开始
    sheet1 = wb.create_sheet('肥仔', index=1)
    sheet1.title = '肥仔1'
    sheet1['B2'] = '肥仔肥仔'
    # print(wb.sheetnames)
    wb.save('fat_boy.xlsx')

    # 打开获取数据
    wb1 = load_workbook('fat_boy.xlsx')
    print(wb1.sheetnames)
    sheet2 = wb.get_sheet_by_name('fat_boy')
    # 第C列
    print(sheet2['C'])
    # 第四行
    print(sheet2['4'])
    # 第C3格的值
    print(sheet["C3"].value)
    # 最大行数
    print(sheet.max_row)
    # 最大列数
    print(sheet.max_column)
    # 所有A列中的值
    for i in sheet["A"]:
        print(i.value, end=" ")
